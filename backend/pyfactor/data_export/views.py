import io
import json
import csv
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from inventory.models import Product
from crm.models import Customer
from sales.models import Invoice, InvoiceItem
from purchases.models import Bill, BillItem, Vendor
from hr.models import Employee
from finance.models import Account
from users.models import UserProfile


class DataExportView(View):
    """Handle data export requests with proper tenant isolation"""
    
    @method_decorator(csrf_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def post(self, request):
        """Export data based on selected types and format"""
        try:
            # Debug logging
            print(f"[DataExportView] Request headers: {dict(request.headers)}")
            print(f"[DataExportView] Request cookies: {request.COOKIES}")
            print(f"[DataExportView] Has user: {hasattr(request, 'user')}")
            print(f"[DataExportView] User authenticated: {request.user.is_authenticated if hasattr(request, 'user') else 'No user'}")
            
            # Check if user is authenticated (middleware should have set this)
            if not hasattr(request, 'user') or not request.user or not request.user.is_authenticated:
                return JsonResponse({'error': 'User not authenticated'}, status=401)
            
            user = request.user
            
            # Get user profile for tenant_id
            try:
                profile = UserProfile.objects.get(user=user)
                tenant_id = profile.tenant_id
            except UserProfile.DoesNotExist:
                return JsonResponse({'error': 'User profile not found'}, status=404)
            
            # Check RBAC permissions
            if profile.role not in ['OWNER', 'ADMIN']:
                return JsonResponse({'error': 'Insufficient permissions. Only OWNER or ADMIN can export data.'}, status=403)
            
            # Parse request body
            body = json.loads(request.body)
            data_types = body.get('dataTypes', [])
            export_format = body.get('format', 'excel')
            date_range = body.get('dateRange', 'all')
            custom_date_range = body.get('customDateRange', {})
            
            print(f"[DataExportView] Processing export for tenant {tenant_id}")
            print(f"[DataExportView] Data types: {data_types}")
            print(f"[DataExportView] Format: {export_format}")
            
            # Collect data for each type
            export_data = {}
            
            for data_type in data_types:
                if data_type == 'products':
                    queryset = Product.objects.filter(tenant_id=tenant_id)
                    data = list(queryset.values(
                        'name', 'sku', 'description', 'unit_price', 'cost_price',
                        'category', 'quantity_on_hand', 'reorder_level', 'tax_rate',
                        'barcode', 'supplier', 'location', 'is_active'
                    ))
                    export_data['Products'] = data
                    
                elif data_type == 'services':
                    queryset = Product.objects.filter(tenant_id=tenant_id, product_type='SERVICE')
                    data = list(queryset.values(
                        'name', 'sku', 'description', 'unit_price', 
                        'category', 'tax_rate', 'is_active'
                    ))
                    export_data['Services'] = data
                    
                elif data_type == 'customers':
                    queryset = Customer.objects.filter(tenant_id=tenant_id)
                    data = list(queryset.values(
                        'name', 'email', 'phone', 'company_name',
                        'address_line1', 'address_line2', 'city', 'state',
                        'postal_code', 'country', 'credit_limit', 'is_active'
                    ))
                    export_data['Customers'] = data
                    
                elif data_type == 'invoices':
                    queryset = Invoice.objects.filter(tenant_id=tenant_id)
                    
                    # Apply date filters
                    if date_range != 'all':
                        queryset = self._apply_date_filter(queryset, date_range, custom_date_range)
                    
                    data = []
                    for invoice in queryset:
                        invoice_data = {
                            'invoice_number': invoice.invoice_number,
                            'customer': invoice.customer.name if invoice.customer else '',
                            'date': invoice.date.strftime('%Y-%m-%d'),
                            'due_date': invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
                            'subtotal': float(invoice.subtotal),
                            'tax_amount': float(invoice.tax_amount),
                            'total': float(invoice.total),
                            'paid_amount': float(invoice.paid_amount),
                            'balance': float(invoice.balance),
                            'status': invoice.status,
                        }
                        data.append(invoice_data)
                    export_data['Invoices'] = data
                    
                elif data_type == 'bills':
                    queryset = Bill.objects.filter(tenant_id=tenant_id)
                    
                    # Apply date filters
                    if date_range != 'all':
                        queryset = self._apply_date_filter(queryset, date_range, custom_date_range)
                    
                    data = []
                    for bill in queryset:
                        bill_data = {
                            'bill_number': bill.bill_number,
                            'vendor': bill.vendor.name if bill.vendor else '',
                            'date': bill.date.strftime('%Y-%m-%d'),
                            'due_date': bill.due_date.strftime('%Y-%m-%d') if bill.due_date else '',
                            'subtotal': float(bill.subtotal),
                            'tax_amount': float(bill.tax_amount),
                            'total': float(bill.total),
                            'paid_amount': float(bill.paid_amount),
                            'balance': float(bill.balance),
                            'status': bill.status,
                        }
                        data.append(bill_data)
                    export_data['Bills'] = data
                    
                elif data_type == 'vendors':
                    queryset = Vendor.objects.filter(tenant_id=tenant_id)
                    data = list(queryset.values(
                        'name', 'email', 'phone', 'company_name',
                        'address_line1', 'address_line2', 'city', 'state',
                        'postal_code', 'country', 'payment_terms', 'is_active'
                    ))
                    export_data['Vendors'] = data
                    
                elif data_type == 'employees':
                    queryset = Employee.objects.filter(tenant_id=tenant_id)
                    data = list(queryset.values(
                        'employee_id', 'first_name', 'last_name', 'email',
                        'phone', 'department', 'position', 'hire_date',
                        'employment_status', 'pay_rate', 'pay_frequency'
                    ))
                    export_data['Employees'] = data
                    
                elif data_type == 'tax-rates':
                    # Tax rates are stored differently in this system
                    # For now, skip tax rates export
                    export_data['Tax Rates'] = []
                    
                elif data_type == 'chart-of-accounts':
                    queryset = Account.objects.filter(tenant_id=tenant_id)
                    data = list(queryset.values(
                        'code', 'name', 'account_type', 'description',
                        'parent_account', 'is_active', 'balance'
                    ))
                    export_data['Chart of Accounts'] = data
            
            # Generate export file based on format
            if export_format == 'excel' and HAS_OPENPYXL:
                response = self._generate_excel_file(export_data)
            elif export_format in ['csv', 'excel']:
                # For CSV, export only the first data type
                first_key = list(export_data.keys())[0] if export_data else None
                if first_key:
                    response = self._generate_csv_file(export_data[first_key], first_key)
                else:
                    return JsonResponse({'error': 'No data to export'}, status=400)
            else:
                return JsonResponse({'error': f'Format {export_format} not supported'}, status=400)
            
            # Set filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if export_format == 'excel' and HAS_OPENPYXL:
                filename = f"dott_export_{timestamp}.xlsx"
            else:
                filename = f"dott_export_{timestamp}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            print(f"[DataExportView] Error: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    def _apply_date_filter(self, queryset, date_range, custom_date_range):
        """Apply date filters to queryset"""
        now = datetime.now()
        
        if date_range == 'thisMonth':
            start_date = datetime(now.year, now.month, 1)
            queryset = queryset.filter(date__gte=start_date)
        elif date_range == 'lastMonth':
            if now.month == 1:
                start_date = datetime(now.year - 1, 12, 1)
                end_date = datetime(now.year - 1, 12, 31)
            else:
                start_date = datetime(now.year, now.month - 1, 1)
                end_date = datetime(now.year, now.month, 1) - timedelta(days=1)
            queryset = queryset.filter(date__gte=start_date, date__lte=end_date)
        elif date_range == 'thisQuarter':
            quarter_month = ((now.month - 1) // 3) * 3 + 1
            start_date = datetime(now.year, quarter_month, 1)
            queryset = queryset.filter(date__gte=start_date)
        elif date_range == 'thisYear':
            start_date = datetime(now.year, 1, 1)
            queryset = queryset.filter(date__gte=start_date)
        elif date_range == 'custom':
            if custom_date_range.get('start'):
                queryset = queryset.filter(date__gte=custom_date_range['start'])
            if custom_date_range.get('end'):
                queryset = queryset.filter(date__lte=custom_date_range['end'])
        
        return queryset
    
    def _generate_csv_file(self, data, name):
        """Generate CSV file"""
        output = io.StringIO()
        
        if data and len(data) > 0:
            # Get headers from first row
            headers = list(data[0].keys())
            
            # Write CSV
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        return response
    
    def _generate_excel_file(self, export_data):
        """Generate Excel file with multiple sheets"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create a sheet for each data type
        for sheet_name, data in export_data.items():
            if data and len(data) > 0:
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet names max 31 chars
                
                # Get headers from first row
                headers = list(data[0].keys())
                
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                
                # Write data
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        # Convert datetime objects to string
                        if hasattr(value, 'strftime'):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        return response